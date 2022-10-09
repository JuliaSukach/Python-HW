# Прочитать сохраненный csv-файл и сохранить данные в excel-файл, кроме возраста - столбец с этими данными не нужен
import csv
import openpyxl

wb = openpyxl.Workbook()

sheet = wb.create_sheet(title='test', index=0)

remove_column_name = 'age'
eliminate_index = -1

with open("data.csv", encoding='utf-8') as r_file:
    file_reader = csv.reader(r_file, delimiter=",")
    for i, row in enumerate(file_reader, start=1):
        column_count = 1
        for j, value in enumerate(row, start=1):
            if value == remove_column_name or j == eliminate_index:
                if eliminate_index < 0:
                    eliminate_index = j
                continue
            c = sheet.cell(row=i, column=column_count)
            c.value = value
            column_count += 1

wb.save("data.xlsx")
